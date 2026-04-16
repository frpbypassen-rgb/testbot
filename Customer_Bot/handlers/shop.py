from config import customer_bot, users, stock, transactions, db
from telebot import types, apihelper
import datetime
import pandas as pd
import io
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import admin_bot, ADMIN_ID

# تعريف جداول قاعدة البيانات الخاصة بالصور والعدادات
db_images = db.section_images
counters = db.counters

# --- 1. القائمة الرئيسية للمتجر (واجهة الصور) ---
@customer_bot.message_handler(func=lambda m: m.text == "🛒 شراء كروت")
def shop_visual_menu(msg):
    uid = msg.chat.id
    user = users.find_one({"_id": uid})
    
    if not user or user.get("status") != "active":
        return customer_bot.send_message(uid, "⚠️ عذراً، يجب تفعيل حسابك من قبل الإدارة أولاً لتتمكن من الشراء.")

    available_categories = stock.distinct("category", {"sold": False})
    
    if not available_categories:
        return customer_bot.send_message(uid, "📭 المتجر فارغ حالياً. يرجى الانتظار لحين إضافة مخزون جديد.")

    customer_bot.send_message(uid, "📸 **مرحباً بك في متجر شركة الأهرام الدولية:**\nتفضل بتصفح الأقسام المتوفرة بالأسفل:", parse_mode="Markdown")

    db_images_list = list(db_images.find({}))
    images_dict = {img['category']: img['file_id'] for img in db_images_list if 'file_id' in img}

    kb_fallback = types.InlineKeyboardMarkup(row_width=2)
    has_images = False

    for cat in available_categories:
        callback_val = f"m_cat_{cat}"
        
        if cat in images_dict:
            try:
                kb_cat = types.InlineKeyboardMarkup()
                kb_cat.add(types.InlineKeyboardButton(f"🔗 دخول قسم {cat}", callback_data=callback_val))
                customer_bot.send_photo(uid, images_dict[cat], caption=f"✨ **قسم: {cat}**", reply_markup=kb_cat, parse_mode="Markdown")
                has_images = True
            except Exception as e:
                print(f"⚠️ خطأ في صورة {cat}: {e}")
                kb_fallback.add(types.InlineKeyboardButton(f"📂 {cat}", callback_data=callback_val))
        else:
            kb_fallback.add(types.InlineKeyboardButton(f"📂 {cat}", callback_data=callback_val))

    if len(kb_fallback.keyboard) > 0:
        text = "🔽 **أقسام إضافية متوفرة:**" if has_images else "🔗 **الأقسام المتوفرة حالياً:**"
        customer_bot.send_message(uid, text, reply_markup=kb_fallback, parse_mode="Markdown")

# --- 2. الدخول للقسم ---
@customer_bot.callback_query_handler(func=lambda call: call.data.startswith("m_cat_"))
def view_subcategories(call):
    print(f"🔘 تم الضغط على قسم: {call.data}")
    
    category_name = call.data.replace("m_cat_", "")
    available_subs = stock.distinct("product", {"category": category_name, "sold": False})
    
    if not available_subs:
        customer_bot.answer_callback_query(call.id, "❌ القسم فارغ حالياً.", show_alert=True)
        return

    kb = types.InlineKeyboardMarkup(row_width=2)
    for sub in available_subs:
        kb.add(types.InlineKeyboardButton(sub, callback_data=f"v_sub_{sub}"))
    
    kb.add(types.InlineKeyboardButton("⬅️ رجوع لكافة الأقسام", callback_data="back_to_shop"))
    
    text = f"📦 **منتجات قسم {category_name}:**\nاختر المنتج المطلوب:"
    
    try:
        customer_bot.delete_message(call.message.chat.id, call.message.message_id)
        customer_bot.send_message(call.message.chat.id, text, reply_markup=kb, parse_mode="Markdown")
        customer_bot.answer_callback_query(call.id)
    except Exception as e:
        try:
            customer_bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=kb, parse_mode="Markdown")
            customer_bot.answer_callback_query(call.id)
        except:
            print(f"🚨 خطأ فني في التحديث: {e}")

def show_product_details(msg, product_id):
    uid = msg.chat.id
    user = users.find_one({"_id": uid})
    user_level = user.get("level", 1) 
    
    product = db.products.find_one({"_id": product_id})
    if not product: return
    
    if user_level == 3:
        final_price = product.get("price_3")
        level_name = "الذهبي 🏆"
    elif user_level == 2:
        final_price = product.get("price_2")
        level_name = "الفضي 🥈"
    else:
        final_price = product.get("price_1")
        level_name = "البرونزي 🥉"

    text = (
        f"📦 **المنتج:** {product.get('name', 'غير معروف')}\n"
        f"⭐ **مستوى حسابك:** {level_name}\n"
        f"💰 **السعر الخاص بك:** `{final_price:.2f}` د.ل"
    )
    customer_bot.send_message(uid, text, parse_mode="Markdown")

# --- 3. طلب الكمية وتحديد السعر حسب مستوى العميل ---
@customer_bot.callback_query_handler(func=lambda call: call.data.startswith("v_sub_"))
def ask_quantity(call):
    uid = call.message.chat.id
    sub_name = call.data.replace("v_sub_", "")
    
    sample = stock.find_one({"product": sub_name, "sold": False})
    
    if not sample:
        customer_bot.answer_callback_query(call.id, "❌ نفذت الكمية!", show_alert=True)
        return

    user = users.find_one({"_id": uid})
    user_level = user.get("level", 1) 
    
    if user_level == 3:
        price = sample.get("price_3", sample.get("price_1", 0.0))
        level_name = "موزع 🥇"
    elif user_level == 2:
        price = sample.get("price_2", sample.get("price_1", 0.0))
        level_name = "جملة 🥈"
    else:
        price = sample.get("price_1", 0.0)
        level_name = "عادي 🥉"

    qty_available = stock.count_documents({"product": sub_name, "sold": False})
    
    text = (
        f"💎 **المنتج:** {sub_name}\n"
        f"⭐ **مستوى حسابك:** {level_name}\n"
        f"💰 **السعر الخاص بك:** `{price:.2f}` د.ل\n"
        f"🔢 **المتوفر:** `{qty_available}` كارت\n\n"
        "✍️ **أرسل الآن الكمية المطلوبة (أرقام فقط):**"
    )
    
    customer_bot.edit_message_text(text, call.message.chat.id, call.message.message_id, parse_mode="Markdown")
    customer_bot.answer_callback_query(call.id)
    customer_bot.register_next_step_handler(call.message, process_final_purchase, sub_name, price, qty_available)

# --- 4. العودة للمتجر ---
@customer_bot.callback_query_handler(func=lambda call: call.data == "back_to_shop")
def back_to_shop(call):
    customer_bot.delete_message(call.message.chat.id, call.message.message_id)
    shop_visual_menu(call.message)
    customer_bot.answer_callback_query(call.id)

# --- 5. إتمام الشراء، الخصم، وتوليد الفاتورة الاحترافية ---
def process_final_purchase(msg, sub_name, price, max_available):
    uid = msg.chat.id
    user = users.find_one({"_id": uid})
    
    try:
        qty = int(msg.text.strip())
    except:
        return customer_bot.send_message(uid, "❌ خطأ: يرجى إرسال الكمية كأرقام فقط. ابدأ من جديد.")

    if qty <= 0 or qty > max_available:
        return customer_bot.send_message(uid, f"❌ الكمية المطلوبة غير متاحة. المتاح {max_available} فقط.")

    total_price = qty * price
    if user.get("balance", 0.0) < total_price:
        return customer_bot.send_message(uid, f"❌ رصيدك لا يكفي. المطلوب: `{total_price:.2f}` د.ل")

    counter_doc = counters.find_one_and_update(
        {"_id": "invoice_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    serial_no = str(counter_doc.get("seq", 1)).zfill(5)

    cards = list(stock.find({"product": sub_name, "sold": False}).limit(qty))
    c_ids = [c["_id"] for c in cards]
    stock.update_many({"_id": {"$in": c_ids}}, {"$set": {"sold": True, "sold_to": uid, "sold_date": datetime.datetime.now()}})
    users.update_one({"_id": uid}, {"$inc": {"balance": -total_price}})
    
    order_id = datetime.datetime.now().strftime("%y%m%d%H%M%S")
    transactions.insert_one({
        "order_id": order_id, "serial_no": serial_no, "user_id": uid, 
        "product": sub_name, "qty": qty, "total_price": total_price, "date": datetime.datetime.now()
    })

    df = pd.DataFrame(cards)
    
    expected_cols = ['product', 'code', 'serial', 'pin', 'op_code']
    for col in expected_cols:
        if col not in df.columns:
            df[col] = pd.NA
            
    df = df[expected_cols]
    df.rename(columns={'product': 'المنتج', 'code': 'الكود', 'serial': 'السيريال', 'pin': 'الرقم السري', 'op_code': 'أوبريشن'}, inplace=True)

    df.replace(r'^\s*$', pd.NA, regex=True, inplace=True)
    df.replace([None, 'nan', 'NaN', ''], pd.NA, inplace=True)
    df.dropna(axis=1, how='all', inplace=True)
    df.fillna("-", inplace=True)

    num_cols = len(df.columns)
    last_col_letter = get_column_letter(num_cols) 

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='فاتورة الأهرام', startrow=5)
        ws = writer.sheets['فاتورة الأهرام']
        ws.sheet_view.rightToLeft = True

        company_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid") 
        password_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f'A1:{last_col_letter}1')
        company_cell = ws['A1']
        company_cell.value = "شركة الأهرام للخدمات الدولية"
        company_cell.fill = company_fill
        company_cell.font = Font(size=20, bold=True, color="000000")
        company_cell.alignment = align

        ws['A2'] = "رقم الفاتورة:"
        ws['B2'] = f"INV-{serial_no}"
        ws['A3'] = "اسم العميل:"
        ws['B3'] = user.get('name', 'عميل الأهرام')
        
        col_before_last = get_column_letter(num_cols - 1 if num_cols > 1 else 1)
        ws[f'{col_before_last}2'] = "التاريخ:"
        ws[f'{last_col_letter}2'] = datetime.datetime.now().strftime("%Y-%m-%d")
        
        for r in range(2, 4):
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).font = Font(bold=True)

        pass_col_idx = None
        for i, col in enumerate(df.columns, 1):
            if col == 'الرقم السري':
                pass_col_idx = i

        for col_num in range(1, num_cols + 1):
            cell = ws.cell(row=6, column=col_num)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_num)].width = 25

        for row_idx in range(7, 7 + len(df)):
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = align
                cell.border = border
                cell.font = Font(bold=True, size=11)
                
                if col_idx == pass_col_idx:
                    cell.fill = password_fill

    output.seek(0)
    
    caption = (
        f"✅ **تمت عملية الشراء بنجاح!**\n"
        f"📑 رقم الفاتورة: `INV-{serial_no}`\n"
        f"💰 القيمة المخصومة: `{total_price:.2f}` د.ل\n"
        f"📂 تجد الكروت المشتراة في الفاتورة المرفقة."
    )
    
    customer_bot.send_document(
        uid, 
        output, 
        visible_file_name=f"AlAhram_INV_{serial_no}.xlsx", 
        caption=caption, 
        parse_mode="Markdown"
    )
    
    try:
        output.seek(0)
        admin_alert = (
            "🚨 **تنبيه مبيعات عاجل - شركة الأهرام**\n"
            "━━━━━━━━━━━━━━━\n"
            f"🛒 **عملية شراء جديدة!**\n"
            f"👤 **العميل:** `{uid}`\n"
            f"📦 **المنتج:** `{sub_name}`\n"
            f"💰 **القيمة:** `{total_price:.2f}` د.ل\n"
            f"🧾 **رقم الفاتورة:** `{serial_no}`\n"
            f"📅 **الوقت:** {datetime.datetime.now().strftime('%H:%M:%S')}\n"
            "━━━━━━━━━━━━━━━\n"
            "📌 *مرفق نسخة من الملف الذي استلمه العميل.*"
        )
        admin_bot.send_message(ADMIN_ID, admin_alert, parse_mode="Markdown")
        admin_bot.send_document(
            ADMIN_ID, 
            output, 
            visible_file_name=f"AlAhram_Copy_{serial_no}.xlsx",
            caption=f"📄 نسخة أرشيفية للفاتورة `{serial_no}`"
        )
    except Exception as e:
        print(f"🚨 فشل إرسال النسخة الإدارية: {e}")