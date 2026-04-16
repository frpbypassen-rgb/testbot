import datetime
import pandas as pd
import io
from config import admin_bot, stock
from telebot import types
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- 1. توليد ملف إكسيل موحد لتعديل الأسعار ---
@admin_bot.message_handler(func=lambda m: m.text == "🏷️ تعديل أسعار المنتجات")
def export_prices_excel(msg):
    uid = msg.chat.id
    status_msg = admin_bot.send_message(uid, "⏳ **جاري تحضير قائمة الأسعار الموحدة...**", parse_mode="Markdown")

    try:
        # جلب المنتجات الفريدة مع أسعارها وترتيبها أبجدياً
        pipeline = [
            {"$match": {"sold": False}},
            {"$group": {
                "_id": "$product",
                "p1": {"$first": "$price_1"},
                "p2": {"$first": "$price_2"},
                "p3": {"$first": "$price_3"}
            }},
            {"$sort": {"_id": 1}} # الترتيب الأبجدي
        ]
        products = list(stock.aggregate(pipeline))

        if not products:
            return admin_bot.edit_message_text("❌ لا يوجد مخزون متاح حالياً لتعديله.", uid, status_msg.message_id)

        output = io.BytesIO()
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M")

        # تجهيز البيانات في قائمة واحدة
        data_list = []
        for item in products:
            data_list.append({
                "الاسم الحالي (لا تعدله)": item['_id'],
                "الاسم الجديد للمنتج": item['_id'],
                "سعر العادي": item['p1'],
                "سعر الجملة": item['p2'],
                "سعر الموزع": item['p3']
            })

        df = pd.DataFrame(data_list)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="تعديل الأسعار", index=False, startrow=4)
            ws = writer.sheets["تعديل الأسعار"]
            ws.sheet_view.rightToLeft = True

            # التنسيق الاحترافي (براند شركة الأهرام)
            brand_blue = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            gold_accent = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            # الهيدر الرئيسي
            ws.merge_cells('A1:E1')
            ws['A1'] = "شركة الأهرام للاتصالات والتقنية - إدارة التسعير"
            ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
            ws['A1'].fill = brand_blue
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40

            # سطر التاريخ والتعليمات
            ws.merge_cells('A2:E2')
            ws['A2'] = f"تاريخ التحديث: {date_str} {time_str} | ملاحظة: عدل الأسعار أو الأسماء ثم ارفع الملف للبوت."
            ws['A2'].alignment = Alignment(horizontal="center")
            ws['A2'].font = Font(bold=True, color="595959")

            # تنسيق رؤوس الجدول
            for col in range(1, 6):
                cell = ws.cell(row=5, column=col)
                cell.fill = gold_accent
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col)].width = 25

        output.seek(0)
        admin_bot.send_document(
            uid, output, 
            visible_file_name=f"AlAhram_PriceList_{date_str}.xlsx",
            caption="✅ **تم استخراج قائمة الأسعار الموحدة.**\n\nقم بتعديل الأسماء أو الأسعار في هذا الملف، ثم أعد إرساله هنا لاعتماد التغييرات.",
            parse_mode="Markdown"
        )
        admin_bot.delete_message(uid, status_msg.message_id)
        admin_bot.register_next_step_handler_by_chat_id(uid, process_price_update_file)

    except Exception as e:
        print(f"🚨 خطأ: {e}")
        admin_bot.send_message(uid, "❌ حدث خطأ فني أثناء تحضير القائمة.")

# --- 2. معالجة الملف المرفوع وتحديث قاعدة البيانات ---
def process_price_update_file(msg):
    uid = msg.chat.id
    if not msg.document:
        return admin_bot.send_message(uid, "❌ تم إلغاء العملية. يجب إرسال ملف الإكسيل.")

    status_wait = admin_bot.send_message(uid, "⏳ جاري فحص البيانات وتحديث المتجر...")

    try:
        file_info = admin_bot.get_file(msg.document.file_id)
        downloaded_file = admin_bot.download_file(file_info.file_path)
        
        df = pd.read_excel(io.BytesIO(downloaded_file), skiprows=4)
        if df.empty:
            return admin_bot.send_message(uid, "❌ الملف فارغ!")

        total_updated_cards = 0
        total_products = 0

        for index, row in df.iterrows():
            old_name = str(row['الاسم الحالي (لا تعدله)']).strip()
            new_name = str(row['الاسم الجديد للمنتج']).strip()
            
            # تحويل آمن للأرقام
            p1 = float(row['سعر العادي'])
            p2 = float(row['سعر الجملة'])
            p3 = float(row['سعر الموزع'])

            # تحديث جماعي لكافة الكروت غير المباعة لهذا المنتج
            res = stock.update_many(
                {"product": old_name, "sold": False},
                {"$set": {
                    "product": new_name,
                    "price_1": p1,
                    "price_2": p2,
                    "price_3": p3
                }}
            )
            if res.modified_count > 0:
                total_updated_cards += res.modified_count
                total_products += 1

        success_text = (
            f"✅ **تم تحديث الأسعار والأسماء بنجاح!**\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📦 عدد المنتجات المعدلة: `{total_products}`\n"
            f"💳 إجمالي الكروت المتأثرة: `{total_updated_cards}`\n"
            f"━━━━━━━━━━━━━━━\n"
            f"✨ المتجر الآن يعرض البيانات الجديدة للعملاء."
        )
        admin_bot.delete_message(uid, status_wait.message_id)
        admin_bot.send_message(uid, success_text, parse_mode="Markdown")

    except Exception as e:
        print(f"🚨 خطأ: {e}")
        admin_bot.send_message(uid, "❌ حدث خطأ أثناء معالجة الملف. تأكد من صحة البيانات والأرقام.")