import datetime
import pandas as pd
import io
from config import admin_bot, stock
from telebot import types
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- 1. عرض ملخص المخزون والزر ---
@admin_bot.message_handler(func=lambda m: m.text == "📦 المخزون")
def show_inventory_summary(msg):
    uid = msg.chat.id
    
    # تجميع الكروت غير المباعة حسب القسم والمنتج لمعرفة الأعداد
    pipeline = [
        {"$match": {"sold": False}},
        {"$group": {
            "_id": {"category": "$category", "product": "$product"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id.category": 1, "_id.product": 1}}
    ]
    
    inventory_summary = list(stock.aggregate(pipeline))
    
    if not inventory_summary:
        return admin_bot.send_message(uid, "📭 **المخزون فارغ تماماً.** لا يوجد أي كروت متاحة للبيع حالياً.", parse_mode="Markdown")
    
    # بناء رسالة الملخص النصية
    total_cards = 0
    text_summary = "📊 **ملخص المخزون الحالي (الكروت المتاحة):**\n━━━━━━━━━━━━━━━\n"
    
    current_category = None
    for item in inventory_summary:
        cat = item["_id"].get("category", "أقسام عامة")
        prod = item["_id"].get("product", "غير معروف")
        count = item["count"]
        total_cards += count
        
        if cat != current_category:
            text_summary += f"\n📂 **{cat}:**\n"
            current_category = cat
            
        text_summary += f"   🔸 {prod}: `{count}` كارت\n"
        
    text_summary += f"\n━━━━━━━━━━━━━━━\n📦 **إجمالي الكروت في المخزن:** `{total_cards}`"

    # زر تحميل الإكسيل
    kb = types.InlineKeyboardMarkup()
    btn_download = types.InlineKeyboardButton("📥 تحميل جرد المخزون الشامل", callback_data="download_full_inventory")
    kb.add(btn_download)
    
    admin_bot.send_message(uid, text_summary, reply_markup=kb, parse_mode="Markdown")

# --- 2. محرك تصدير إكسيل المخزون (بدون أخطاء + تنسيق احترافي) ---
@admin_bot.callback_query_handler(func=lambda call: call.data == "download_full_inventory")
def export_inventory_excel(call):
    uid = call.message.chat.id
    admin_bot.answer_callback_query(call.id, "⏳ جاري تجميع الكروت وتجهيز الملف...")
    status_msg = admin_bot.send_message(uid, "⏳ **جاري بناء ملف الجرد وتنسيقه...**", parse_mode="Markdown")

    try:
        available_cards = list(stock.find({"sold": False}))
        
        if not available_cards:
            return admin_bot.edit_message_text("❌ لم يتم العثور على كروت متاحة للجرد.", uid, status_msg.message_id)

        output = io.BytesIO()
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        categories = set([c.get('category', 'أقسام عامة') for c in available_cards])
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for cat in categories:
                cat_cards = [c for c in available_cards if c.get('category', 'أقسام عامة') == cat]
                
                # تجهيز البيانات كـ نصوص (Strings) لتجنب أخطاء Pandas مع الخلايا الفارغة
                data_list = []
                for card in cat_cards:
                    # التحقق الآمن من التاريخ
                    add_date = card.get("added_date")
                    date_val = add_date.strftime("%Y-%m-%d %H:%M") if hasattr(add_date, "strftime") else "-"
                    
                    data_list.append({
                        "اسم المنتج": str(card.get("product", card.get("subcategory", "-")) or "-"),
                        "السعر الأساسي": f"{card.get('price_1', 0.0):.2f} د.ل",
                        "رقم الكارت (Code)": str(card.get("code", "-") or "-"),
                        "الرقم التسلسلي (Serial)": str(card.get("serial", "-") or "-"),
                        "الرقم السري (Pin)": str(card.get("pin", "-") or "-"),
                        "أوبريشن كود": str(card.get("op_code", "-") or "-"),
                        "تاريخ الإضافة": date_val
                    })
                
                df = pd.DataFrame(data_list)
                
                safe_sheet_name = "".join([c for c in str(cat) if c.isalnum() or c in " _-"])[:31]
                if not safe_sheet_name: safe_sheet_name = "Sheet"
                
                # كتابة البيانات مع ترك 4 سطور فارغة في الأعلى للهيدر
                df.to_excel(writer, sheet_name=safe_sheet_name, index=False, startrow=4)
                
                # --- التنسيق الجمالي (Openpyxl) ---
                ws = writer.sheets[safe_sheet_name]
                ws.sheet_view.rightToLeft = True # دعم العربي
                
                brand_blue = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                gold_accent = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                
                num_cols = len(df.columns)
                last_col = get_column_letter(num_cols)
                
                # 1. هيدر الشركة الرئيسي
                ws.merge_cells(f'A1:{last_col}1')
                ws['A1'] = "شركة الأهرام للاتصالات والتقنية - تقرير جرد المخزون"
                ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
                ws['A1'].fill = brand_blue
                ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 35
                
                # 2. سطر الوقت والتاريخ والإحصائيات
                ws.merge_cells(f'A2:{last_col}2')
                ws['A2'] = f"📅 التاريخ: {date_str} | ⏰ الوقت: {time_str} | 📂 القسم: {cat} | 📦 عدد الكروت المتاحة: {len(cat_cards)}"
                ws['A2'].font = Font(size=12, bold=True, color="008000")
                ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
                
                # 3. تنسيق عناوين الأعمدة وتوسيعها
                for i in range(1, num_cols + 1):
                    # تلوين العناوين في السطر الخامس
                    cell = ws.cell(row=5, column=i)
                    cell.fill = gold_accent
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    # توسيع العمود
                    ws.column_dimensions[get_column_letter(i)].width = 25

        output.seek(0)
        file_name = f"AlAhram_Inventory_{date_str}.xlsx"
        
        caption = (
            f"✅ **تم استخراج جرد المخزون بنجاح!**\n"
            f"📅 التاريخ: `{date_str}`\n"
            f"⏰ الوقت: `{time_str}`\n"
            f"📂 الملف منسق ومقسم إلى شيتات حسب الأقسام."
        )
        
        admin_bot.send_document(
            uid, output, 
            visible_file_name=file_name,
            caption=caption, parse_mode="Markdown"
        )
        admin_bot.delete_message(uid, status_msg.message_id)

    except Exception as e:
        print(f"🚨 خطأ في استخراج المخزون: {e}")
        admin_bot.send_message(uid, "❌ حدث خطأ داخلي أثناء معالجة الجرد.")