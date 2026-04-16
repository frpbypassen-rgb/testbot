import datetime
import pandas as pd
import io
from config import admin_bot, users, transactions, db
from telebot import types
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter # 👈 الاستدعاء المضاف لتوسيع الأعمدة بأمان

# --- المحرك الرئيسي لاستخراج التقرير الشامل ---
@admin_bot.message_handler(func=lambda m: m.text == "📂 قائمة المشتركين")
def export_subscribers_excel(msg):
    uid = msg.chat.id
    status_msg = admin_bot.send_message(uid, "⏳ **جاري إنشاء التقرير الرسمي لشركة الأهرام...**\nيرجى الانتظار، يتم الآن تنسيق البيانات احترافياً.", parse_mode="Markdown")

    try:
        all_customers = list(users.find({}))
        if not all_customers:
            return admin_bot.edit_message_text("❌ لا يوجد مستخدمين مسجلين حالياً.", uid, status_msg.message_id)

        output = io.BytesIO()
        today_date = datetime.datetime.now().strftime("%Y-%m-%d")
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for customer in all_customers:
                c_id = customer.get("_id")
                c_name = customer.get("name", "عميل_بدون_اسم")
                
                # اسم الشيت
                safe_sheet_name = "".join([c for c in str(c_name) if c.isalnum() or c in " "])[:20]
                sheet_name = f"{safe_sheet_name}_{str(c_id)[-4:]}"

                # 1. تجهيز البيانات
                # سجل البروفايل
                profile_df = pd.DataFrame({
                    "الاسم الكامل": [customer.get("name", "-")],
                    "رقم الهاتف": [customer.get("phone", "-")],
                    "رقم الـ ID": [c_id],
                    "تاريخ التسجيل": [customer.get("reg_date", "غير متوفر")],
                    "الرصيد": [f"{customer.get('balance', 0.0):.2f} د.ل"],
                    "المستوى": [customer.get("level", 1)]
                })

                # سجل المشتريات
                user_purchases = list(transactions.find({"user_id": c_id}))
                df_purchases = pd.DataFrame([{
                    "المنتج": tx.get("product", "-"),
                    "الفاتورة": tx.get("serial_no", "-"),
                    "التاريخ": tx.get("date").strftime("%Y-%m-%d") if hasattr(tx.get("date"), "strftime") else "-",
                    "القيمة": f"{tx.get('total_price', 0.0):.2f} د.ل"
                } for tx in user_purchases]) if user_purchases else pd.DataFrame([{"ملاحظة": "لا توجد مشتريات"}])

                # سجل الشكاوي
                user_complaints = list(db.complaints.find({"user_id": c_id}))
                df_complaints = pd.DataFrame([{
                    "الشكوى": comp.get("text", "-"),
                    "الحالة": "تم الحل ✅" if comp.get("status") == "resolved" else "انتظار ⏳",
                    "الرد": comp.get("admin_reply", "-")
                } for comp in user_complaints]) if user_complaints else pd.DataFrame([{"ملاحظة": "لا توجد شكاوي"}])

                # 2. كتابة البيانات في الشيت مع ترك مساحة للهيدر
                profile_df.to_excel(writer, sheet_name=sheet_name, startrow=4, index=False)
                p_end = 4 + len(profile_df) + 2
                df_purchases.to_excel(writer, sheet_name=sheet_name, startrow=p_end, index=False)
                c_start = p_end + len(df_purchases) + 2
                df_complaints.to_excel(writer, sheet_name=sheet_name, startrow=c_start, index=False)

                # --- 3. لمسة التنسيق الاحترافية (Openpyxl) ---
                ws = writer.sheets[sheet_name]
                ws.sheet_view.rightToLeft = True
                
                # تحديد الألوان
                brand_blue = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                light_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                gold_accent = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                
                # أ. هيدر الشركة الرئيسي
                ws.merge_cells('A1:F1')
                title_cell = ws['A1']
                title_cell.value = "شركة الأهرام للاتصالات والتقنية"
                title_cell.fill = brand_blue
                title_cell.font = Font(size=22, bold=True, color="FFFFFF")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 40

                # ب. سطر التاريخ
                ws.merge_cells('A2:F2')
                date_cell = ws['A2']
                date_cell.value = f"تاريخ استخراج التقرير: {today_date}"
                date_cell.font = Font(size=12, italic=True, color="595959")
                date_cell.alignment = Alignment(horizontal="center")

                # ج. تنسيق رؤوس الجداول
                def format_header(row_num, text, color_fill):
                    ws.merge_cells(f'A{row_num}:F{row_num}')
                    cell = ws.cell(row=row_num, column=1)
                    cell.value = text
                    cell.fill = color_fill
                    cell.font = Font(bold=True, size=14)
                    cell.alignment = Alignment(horizontal="right")

                format_header(4, "📋 بيانات العميل الأساسية:", light_blue)
                format_header(p_end, "🛒 سجل المشتريات المالي:", light_blue)
                format_header(c_start, "📩 سجل الشكاوي والدعم الفني:", light_blue)

                # د. تلوين عناوين الأعمدة في الجداول
                for row in [5, p_end+1, c_start+1]:
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell.fill = gold_accent
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")

                # 🟢 التعديل الجذري: ضبط العرض للأعمدة الـ 6 بأمان تام بعيداً عن الدمج
                for i in range(1, 7):
                    ws.column_dimensions[get_column_letter(i)].width = 25

        output.seek(0)
        file_name = f"AlAhram_Report_{today_date}.xlsx"
        
        admin_bot.send_document(
            uid, output, visible_file_name=file_name,
            caption=f"✅ **تم توليد تقرير شركة الأهرام بنجاح**\n📂 الملف منظم وجاهز للطباعة أو المراجعة.",
            parse_mode="Markdown"
        )
        admin_bot.delete_message(uid, status_msg.message_id)

    except Exception as e:
        print(f"🚨 خطأ: {e}")
        admin_bot.send_message(uid, "❌ حدث خطأ أثناء تنسيق الملف.")