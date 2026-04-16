import pandas as pd
import io
import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import admin_bot, transactions, stock, users, customer_bot
from telebot import types

# --- 1. طلب رقم الفاتورة من الأدمن ---
@admin_bot.message_handler(func=lambda m: m.text == "🔍 البحث عن فاتورة")
def ask_invoice_id(msg):
    res = admin_bot.send_message(
        msg.chat.id, 
        "🆔 **الرجاء إرسال رقم الفاتورة (Order ID) للبحث عنها:**\n*(مثال: 00001 أو INV-00001)*", 
        parse_mode="Markdown"
    )
    admin_bot.register_next_step_handler(res, process_invoice_search)

# --- 2. البحث وتوليد الفاتورة الاحترافية من الأرشيف ---
def process_invoice_search(msg):
    raw_input = msg.text.strip()
    
    # تنظيف الرقم للبحث الشامل
    clean_number = raw_input.replace("INV-", "").replace("inv-", "").strip()
    
    # البحث في قاعدة البيانات عن الفاتورة بأي صيغة
    order = transactions.find_one({
        "$or": [
            {"order_id": raw_input},              
            {"serial_no": raw_input},             
            {"order_id": clean_number},           
            {"serial_no": clean_number},          
            {"order_id": f"INV-{clean_number}"},  
            {"serial_no": str(clean_number).zfill(5)} # دعم الأصفار على اليسار
        ]
    })
    
    if not order:
        return admin_bot.send_message(
            msg.chat.id, 
            f"⚠️ **لم يتم العثور على الفاتورة.**\nتأكد من رقم الفاتورة الصحيح.", 
            parse_mode="Markdown"
        )

    admin_bot.send_message(msg.chat.id, "🔄 جاري سحب الأكواد من الأرشيف وبناء الفاتورة...")

    try:
        uid = order.get('user_id')
        product_name = order.get('product')
        qty = order.get('qty', 1)
        serial_no = order.get('serial_no', clean_number)

        # جلب بيانات العميل
        user = users.find_one({"_id": uid})
        customer_name = user.get('name', 'غير مسجل') if user else "حساب محذوف"

        # جلب الأكواد الدقيقة من المخزن التي تم بيعها لهذا العميل
        sold_cards = list(stock.find({
            "sold_to": uid,
            "$or": [{"product": product_name}, {"subcategory": product_name}]
        }).sort([("sold_date", -1)]).limit(qty))

        if not sold_cards:
            return admin_bot.send_message(msg.chat.id, "❌ تم العثور على الفاتورة، لكن الأكواد المرتبطة بها غير موجودة في الأرشيف.")

        # --- بناء الفاتورة الاحترافية المنسقة ---
        df = pd.DataFrame(sold_cards)
        
        expected_cols = ['product', 'code', 'serial', 'pin', 'op_code']
        for col in expected_cols:
            if col not in df.columns:
                if col == 'product' and 'subcategory' in df.columns: df['product'] = df['subcategory']
                elif col == 'product' and 'name' in df.columns: df['product'] = df['name']
                else: df[col] = pd.NA

        df = df[expected_cols]
        df.rename(columns={'product': 'المنتج', 'code': 'الكود', 'serial': 'السيريال', 'pin': 'الرقم السري', 'op_code': 'أوبريشن'}, inplace=True)

        df.replace(r'^\s*$', pd.NA, regex=True, inplace=True)
        df.replace([None, 'nan', 'NaN', ''], pd.NA, inplace=True)
        df.dropna(axis=1, how='all', inplace=True)
        df.fillna("-", inplace=True)

        num_cols = len(df.columns)
        last_col_letter = get_column_letter(num_cols)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='أرشيف الأهرام', startrow=5)
            ws = writer.sheets['أرشيف الأهرام']
            ws.sheet_view.rightToLeft = True

            company_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid") 
            password_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            align = Alignment(horizontal="center", vertical="center")

            ws.merge_cells(f'A1:{last_col_letter}1')
            company_cell = ws['A1']
            company_cell.value = "شركة الأهرام - نسخة أرشيفية"
            company_cell.fill = company_fill
            company_cell.font = Font(size=18, bold=True, color="000000")
            company_cell.alignment = align

            ws['A2'] = "رقم الفاتورة:"
            ws['B2'] = f"INV-{serial_no}"
            ws['A3'] = "اسم العميل:"
            ws['B3'] = customer_name
            
            col_before_last = get_column_letter(num_cols - 1 if num_cols > 1 else 1)
            ws[f'{col_before_last}2'] = "التاريخ:"
            ws[f'{last_col_letter}2'] = order.get('date', datetime.datetime.now()).strftime("%Y-%m-%d")
            
            for r in range(2, 4):
                for c in range(1, num_cols + 1):
                    ws.cell(row=r, column=c).font = Font(bold=True)

            pass_col_idx = None
            for i, col in enumerate(df.columns, 1):
                if col == 'الرقم السري': pass_col_idx = i

            for col_num in range(1, num_cols + 1):
                cell = ws.cell(row=6, column=col_num)
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = align
                cell.border = border
                ws.column_dimensions[get_column_letter(col_num)].width = 25

            for row_idx in range(7, 7 + len(df)):
                for col_idx in range(1, num_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = align
                    cell.border = border
                    cell.font = Font(bold=True, size=11)
                    if col_idx == pass_col_idx: cell.fill = password_fill

        output.seek(0)
        file_name = f"AlAhram_Archive_INV_{serial_no}.xlsx"
        
        summary = (
            f"🧾 **بيانات الفاتورة المكتشفة:**\n"
            f"━━━━━━━━━━━━━━━\n"
            f"👤 **العميل:** {customer_name}\n"
            f"📦 **المنتج:** {product_name}\n"
            f"🔢 **الكمية:** `{qty}` كروت\n"
            f"💰 **المبلغ:** `{order.get('total_price', 0):.2f}` د.ل\n"
            f"━━━━━━━━━━━━━━━"
        )
        admin_bot.send_message(msg.chat.id, summary, parse_mode="Markdown")
        admin_bot.send_document(msg.chat.id, output, visible_file_name=file_name, caption="📄 ملف الإكسيل مع الأكواد الدقيقة.")

    except Exception as e:
        print(f"🚨 خطأ في استرجاع الفاتورة: {e}")
        admin_bot.send_message(msg.chat.id, f"❌ حدث خطأ داخلي أثناء استرجاع الأكواد.")