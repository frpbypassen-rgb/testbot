import pandas as pd
import io
import datetime
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from config import admin_bot, transactions, users

# --- محرك تقرير المبيعات الشامل ---
@admin_bot.message_handler(func=lambda m: m.text == "📊 تقرير المبيعات")
def generate_general_sales_report(msg):
    uid = msg.chat.id
    status_msg = admin_bot.send_message(uid, "⏳ **جاري تجميع بيانات المبيعات وإنشاء التقرير...**", parse_mode="Markdown")

    try:
        # جلب كل المبيعات وترتيبها من الأحدث للأقدم
        all_tx = list(transactions.find({}).sort("date", -1))
        
        if not all_tx:
            return admin_bot.edit_message_text("📭 لا توجد أي مبيعات مسجلة في النظام حتى الآن.", uid, status_msg.message_id)

        # حساب الإحصائيات العامة
        total_revenue = sum(tx.get("total_price", 0) for tx in all_tx)
        total_items = sum(tx.get("qty", 0) for tx in all_tx)
        total_orders = len(all_tx)

        # تجهيز البيانات للإكسيل
        tx_data = []
        for tx in all_tx:
            user = users.find_one({"_id": tx.get("user_id")})
            customer_name = user.get("name", "غير مسجل") if user else "حساب محذوف"

            tx_data.append({
                "رقم الفاتورة": tx.get("serial_no", tx.get("order_id", "-")),
                "تاريخ العملية": tx.get("date").strftime("%Y-%m-%d %H:%M") if hasattr(tx.get("date"), "strftime") else tx.get("date"),
                "اسم العميل": customer_name,
                "المنتج": tx.get("product", "-"),
                "الكمية المباعة": tx.get("qty", 0),
                "القيمة (د.ل)": tx.get("total_price", 0.0)
            })

        df_tx = pd.DataFrame(tx_data)

        # بناء ملف الإكسيل
        output = io.BytesIO()
        today_date = datetime.datetime.now().strftime("%Y-%m-%d")

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_tx.to_excel(writer, sheet_name="سجل المبيعات", startrow=5, index=False)
            ws = writer.sheets["سجل المبيعات"]
            ws.sheet_view.rightToLeft = True

            # التنسيق الجمالي لشركة الأهرام
            brand_blue = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            gold_accent = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            # الهيدر الرئيسي
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "تقرير مبيعات شركة الأهرام الشامل"
            title_cell.fill = brand_blue
            title_cell.font = Font(size=18, bold=True, color="FFFFFF")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 35

            # سطر الإحصائيات
            ws.merge_cells('A2:F2')
            ws['A2'].value = f"تاريخ التقرير: {today_date} | إجمالي الإيرادات: {total_revenue:.2f} د.ل | إجمالي الكروت: {total_items}"
            ws['A2'].font = Font(size=12, bold=True, color="008000")
            ws['A2'].alignment = Alignment(horizontal="center")

            # تلوين عناوين الأعمدة
            for col in range(1, 7):
                cell = ws.cell(row=6, column=col)
                cell.fill = gold_accent
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # ضبط عرض الأعمدة للجدول
            for i in range(1, 7):
                ws.column_dimensions[get_column_letter(i)].width = 22

        output.seek(0)
        
        # رسالة الملخص للأدمن
        summary = (
            f"📊 **موجز المبيعات السريع:**\n"
            f"━━━━━━━━━━━━━━━\n"
            f"💰 **إجمالي الإيرادات:** `{total_revenue:.2f}` د.ل\n"
            f"📦 **إجمالي الكروت المباعة:** `{total_items}` كارت\n"
            f"🧾 **عدد الفواتير المصدرة:** `{total_orders}` فاتورة\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📥 *التفاصيل الدقيقة والأسماء تجدها في الإكسيل المرفق.*"
        )
        
        admin_bot.send_document(
            uid, output, 
            visible_file_name=f"AlAhram_Sales_Report_{today_date}.xlsx",
            caption=summary, parse_mode="Markdown"
        )
        admin_bot.delete_message(uid, status_msg.message_id)

    except Exception as e:
        print(f"🚨 خطأ في تقرير المبيعات: {e}")
        admin_bot.send_message(uid, "❌ حدث خطأ أثناء توليد التقرير.")