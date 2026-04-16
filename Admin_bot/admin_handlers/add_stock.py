import datetime
import pandas as pd
import io
from config import admin_bot, stock
from telebot import types
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ذاكرة مؤقتة معقدة لحفظ مسار الإدخال
add_stock_state = {}

# ==========================================
# 1. شاشة البداية: اختيار القسم (Category)
# ==========================================
@admin_bot.message_handler(func=lambda m: m.text == "📦 إضافة أكواد")
def start_add_stock(msg):
    uid = msg.chat.id
    add_stock_state[uid] = {} # تصفير الذاكرة
    
    # جلب الأقسام الموجودة مسبقاً من الداتابيز
    existing_categories = stock.distinct("category")
    
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    for cat in existing_categories:
        if cat: kb.add(types.KeyboardButton(cat))
    kb.add(types.KeyboardButton("➕ قسم جديد"), types.KeyboardButton("❌ إلغاء العملية"))
    
    res = admin_bot.send_message(
        uid, 
        "📦 **نظام إدارة المخزون المطور:**\n\nاختر **القسم الرئيسي** من القائمة، أو اضغط على (قسم جديد):", 
        reply_markup=kb, parse_mode="Markdown"
    )
    admin_bot.register_next_step_handler(res, handle_category_selection)

# ==========================================
# 2. معالجة القسم واختيار المنتج (Product)
# ==========================================
def handle_category_selection(msg):
    uid = msg.chat.id
    text = msg.text.strip()
    
    if text == "❌ إلغاء العملية":
        return cancel_operation(uid)
        
    if text == "➕ قسم جديد":
        res = admin_bot.send_message(uid, "📝 أرسل اسم **القسم الجديد** الآن:", reply_markup=types.ReplyKeyboardRemove())
        return admin_bot.register_next_step_handler(res, process_new_category)
        
    process_category(uid, text)

def process_new_category(msg):
    process_category(msg.chat.id, msg.text.strip())

def process_category(uid, category_name):
    add_stock_state[uid]['category'] = category_name
    
    # جلب المنتجات المرتبطة بهذا القسم
    existing_products = stock.distinct("product", {"category": category_name})
    
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    for prod in existing_products:
        if prod: kb.add(types.KeyboardButton(prod))
    kb.add(types.KeyboardButton("➕ منتج جديد"), types.KeyboardButton("❌ إلغاء العملية"))
    
    res = admin_bot.send_message(
        uid, 
        f"📂 القسم: `{category_name}`\nاختر **المنتج** أو أضف منتجاً جديداً:", 
        reply_markup=kb, parse_mode="Markdown"
    )
    admin_bot.register_next_step_handler(res, handle_product_selection)

# ==========================================
# 3. معالجة المنتج وتوليد القالب (Template)
# ==========================================
def handle_product_selection(msg):
    uid = msg.chat.id
    text = msg.text.strip()
    
    if text == "❌ إلغاء العملية": return cancel_operation(uid)
        
    if text == "➕ منتج جديد":
        res = admin_bot.send_message(uid, "📝 أرسل اسم **المنتج الجديد** الآن:", reply_markup=types.ReplyKeyboardRemove())
        return admin_bot.register_next_step_handler(res, process_new_product)
        
    process_product(uid, text)

def process_new_product(msg):
    process_product(msg.chat.id, msg.text.strip())

def process_product(uid, product_name):
    add_stock_state[uid]['product'] = product_name
    
    # إرجاع الكيبورد الأساسي للأدمن
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("📊 تقرير المبيعات", "📩 الشكاوي المعلقة", "📦 إضافة أكواد", "💰 شحن رصيد", "🔍 البحث عن فاتورة", "📂 أرشيف الشكاوي")
    
    # توليد قالب الإكسيل (Template)
    df_template = pd.DataFrame(columns=['code', 'serial', 'pin', 'op_code'])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_template.to_excel(writer, index=False)
    output.seek(0)
    output.name = f"Template_{product_name}.xlsx"
    
    admin_bot.send_document(
        uid, output, 
        caption=f"✅ تم تحديد المنتج: `{product_name}`\n\n"
                f"📥 **الخطوة 1:** قم بتحميل هذا القالب.\n"
                f"📝 **الخطوة 2:** ضع الأكواد داخل عمود `code` (والسيريالات إن وجدت).\n"
                f"📤 **الخطوة 3:** أرسل الملف المعبأ هنا ليتم فحصه واعتماده.",
        reply_markup=kb, parse_mode="Markdown"
    )
    admin_bot.register_next_step_handler_by_chat_id(uid, validate_uploaded_file)

# ==========================================
# 4. استلام الملف وفرز الأكواد (المقبول والمرفوض)
# ==========================================
def validate_uploaded_file(msg):
    uid = msg.chat.id
    
    if msg.text: # لو ضغط على زر بالخطأ
        return admin_bot.send_message(uid, "❌ تم إلغاء الإضافة، يرجى البدء من جديد.")
        
    if not msg.document:
        res = admin_bot.send_message(uid, "❌ يرجى إرسال ملف الإكسيل المعبأ:")
        return admin_bot.register_next_step_handler(res, validate_uploaded_file)

    admin_bot.send_message(uid, "⏳ جاري فحص الملف وفرز الأكواد (مقبول / مرفوض)...")

    try:
        file_info = admin_bot.get_file(msg.document.file_id)
        downloaded_file = admin_bot.download_file(file_info.file_path)
        df = pd.read_excel(io.BytesIO(downloaded_file))

        if df.empty or 'code' not in df.columns:
            res = admin_bot.send_message(uid, "❌ الملف فارغ أو لا يحتوي على عمود `code`. استخدم القالب المرسل وأعد المحاولة:")
            return admin_bot.register_next_step_handler(res, validate_uploaded_file)

        # تنظيف الأكواد
        df = df.dropna(subset=['code'])
        df['code'] = df['code'].astype(str).str.strip()

        # 1. فحص التكرار داخل الملف نفسه
        internal_duplicates_df = df[df.duplicated('code', keep='first')].copy()
        unique_file_df = df.drop_duplicates('code', keep='first').copy()

        # 2. فحص التكرار مع قاعدة البيانات
        codes_list = unique_file_df['code'].tolist()
        duplicates_in_db = list(stock.find({"code": {"$in": codes_list}}))
        dup_db_codes = [d.get('code') for d in duplicates_in_db]

        # فرز المقبول والمرفوض
        accepted_df = unique_file_df[~unique_file_df['code'].isin(dup_db_codes)].copy()
        db_duplicates_df = unique_file_df[unique_file_df['code'].isin(dup_db_codes)].copy()

        # تجميع وتجهيز المرفوضات
        db_duplicates_df['سبب الرفض'] = "موجود مسبقاً في المنظومة"
        internal_duplicates_df['سبب الرفض'] = "مكرر داخل الملف المرفوع"
        rejected_df = pd.concat([db_duplicates_df, internal_duplicates_df])

        # إرسال تقرير المرفوضات (إن وجد)
        if not rejected_df.empty:
            report_rejected = generate_excel_report(rejected_df, "أكواد مرفوضة - شركة الأهرام", "C00000")
            admin_bot.send_document(
                uid, report_rejected, 
                visible_file_name="Rejected_Codes.xlsx", 
                caption=f"❌ تم العثور على `{len(rejected_df)}` كود مكرر (مرفوض)."
            )

        # إذا كانت كل الأكواد مكررة
        if accepted_df.empty:
            return admin_bot.send_message(uid, "❌ تم رفض جميع الأكواد في الملف لأنها مكررة. العملية انتهت.")

        # حفظ البيانات الصالحة في الذاكرة المؤقتة
        add_stock_state[uid]['valid_data'] = accepted_df.to_dict('records')

        # الانتقال للخطوة الأخيرة (الأسعار) للأكواد المقبولة فقط
        res = admin_bot.send_message(
            uid, 
            f"✅ **اجتاز `{len(accepted_df)}` كود الفحص بنجاح!**\n\n"
            "💰 أرسل الآن **الأسعار الثلاثة** مفصولة بمسافة (عادي، جملة، موزع):\n"
            "*(مثال: 10.00 9.50 9.00)*", parse_mode="Markdown"
        )
        admin_bot.register_next_step_handler(res, finalize_stock_insertion)

    except Exception as e:
        print(f"🚨 خطأ الفحص: {e}")
        res = admin_bot.send_message(uid, "❌ حدث خطأ فني أثناء قراءة الملف، تأكد من الصيغة وأعد الإرسال:")
        admin_bot.register_next_step_handler(res, validate_uploaded_file)

# ==========================================
# 5. استلام الأسعار والإدخال النهائي للـ DB
# ==========================================
def finalize_stock_insertion(msg):
    uid = msg.chat.id
    try:
        prices = msg.text.strip().split()
        if len(prices) != 3: raise ValueError("not_three")
            
        p1 = round(float(prices[0]), 2)
        p2 = round(float(prices[1]), 2)
        p3 = round(float(prices[2]), 2)
        
        state = add_stock_state.get(uid, {})
        category = state.get('category')
        product = state.get('product')
        valid_records = state.get('valid_data', [])
        
        stock_to_insert = []
        for row in valid_records:
            stock_item = {
                "category": category,
                "product": product,
                "price_1": p1,
                "price_2": p2,
                "price_3": p3,
                "sold": False,          
                "sold_to": None,        
                "added_date": datetime.datetime.now()
            }
            stock_item.update(row)
            stock_to_insert.append(stock_item)

        if stock_to_insert:
            stock.insert_many(stock_to_insert)
            
            # استخراج تقرير المقبولات الذي تم حفظه
            accepted_df = pd.DataFrame(valid_records)
            report_accepted = generate_excel_report(accepted_df, "أكواد مقبولة - شركة الأهرام", "008000")
            
            success_msg = (
                f"✅ **تم اعتماد وإضافة المخزون!**\n"
                f"━━━━━━━━━━━━━━━\n"
                f"📂 **القسم:** `{category}`\n"
                f"📦 **المنتج:** `{product}`\n"
                f"🔢 **الكمية المضافة:** `{len(stock_to_insert)}` كارت\n"
                f"📊 **الأسعار:** `{p1:.2f}` | `{p2:.2f}` | `{p3:.2f}`\n"
                f"━━━━━━━━━━━━━━━"
            )
            admin_bot.send_message(uid, success_msg, parse_mode="Markdown")
            admin_bot.send_document(
                uid, report_accepted, 
                visible_file_name="Accepted_Codes.xlsx", 
                caption="✅ سجل الأكواد الصالحة والمضافة للمخزون بنجاح"
            )
        
        add_stock_state.pop(uid, None)

    except ValueError as e:
        res = admin_bot.send_message(uid, "❌ خطأ في تنسيق الأسعار. أرسل 3 أرقام مفصولة بمسافة (مثال: 10 9 8):")
        admin_bot.register_next_step_handler(res, finalize_stock_insertion)

# --- دالة مساعدة لتوليد تقارير الإكسيل (المقبول والمرفوض) ---
def generate_excel_report(df, title, header_color):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='التقرير', startrow=4)
        ws = writer.sheets['التقرير']
        ws.sheet_view.rightToLeft = True
        
        fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        
        num_cols = len(df.columns) if len(df.columns) > 0 else 1
        last_col_letter = get_column_letter(num_cols)
        
        ws.merge_cells(f'A1:{last_col_letter}1')
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = fill
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        for col in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
            
    output.seek(0)
    return output

# --- دالة مساعدة لإلغاء العملية ---
def cancel_operation(uid):
    try:
        add_stock_state.pop(uid, None)
    except:
        pass
        
    admin_bot.clear_step_handler_by_chat_id(chat_id=uid)
    
    # بناء لوحة الإدارة الشاملة بطريقة المصفوفة الآمنة
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    buttons = [
        "📊 تقرير المبيعات", "📩 الشكاوي المعلقة",
        "📦 إضافة أكواد", "⚙️ تحكم العملاء",
        "⚙️ إعدادات البوت", "💰 شحن رصيد",
        "💳 أكواد الشحن", "📂 قائمة المشتركين",
        "🖼️ إدارة الصور", "🔍 البحث عن فاتورة",
        "🏷️ تعديل أسعار المنتجات", "📦 المخزون"
    ]
    kb.add(*buttons)
    
    admin_bot.send_message(uid, "🚫 **تم إلغاء العملية والعودة للقائمة الرئيسية.**", reply_markup=kb, parse_mode="Markdown")